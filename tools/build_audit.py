#!/usr/bin/env python3
"""
build_audit.py — turn the extension's hubspot_workflows_<portal>.csv into the
polished FullFunnel workflow-audit .xlsx (6 sheets, branded).

Usage:
    python3 build_audit.py "path/to/hubspot_workflows_545075.csv" [out.xlsx]

Structural + enrollment figures come straight from the CSV (exact). The "Function"
taxonomy and the near-duplicate clustering are heuristic analysis layers — useful
consolidation signal, not a byte-match to any prior hand-built audit.

Reusable across portals: no portal-specific data is baked in.
"""
import csv, sys, re, os, datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCENT = "1F6FEB"           # FullFunnel blue
DARK   = "0B2A4A"
LIGHT  = "EEF3FF"
GREY   = "888888"

CSV_COLS = ["name","status","object_type","flow_type","trigger_type","reenrollment",
            "description","total_enrolled","enrolled_7d","unique_enrolled",
            "currently_enrolled","active_issues","created_in","created_on",
            "created_by","updated_on","updated_by","id","url"]

TRIGGER_LABEL = {"LIST":"List membership","LIST_MEMBERSHIP":"List membership",
                 "EVENT":"Event-based","EVENT_BASED":"Event-based","NEW":"New",
                 "SCHEDULE":"Scheduled","FORM_SUBMISSION":"Form submission"}

# ---- helpers ----------------------------------------------------------------
def num(v):
    try: return int(float(v))
    except: return 0

def is_on(r): return str(r["status"]).strip().upper() == "ON"

def trigger_label(v):
    if not v: return ""
    return TRIGGER_LABEL.get(str(v).strip().upper(), str(v).replace("_"," ").title())

# ---- Function taxonomy (ordered keyword rules) ------------------------------
def classify(r):
    n = (r["name"] or "").lower(); d = (r["description"] or "").lower(); t = n + " " + d
    has = lambda *ks: any(k in t for k in ks)
    if re.search(r"\btest\b", n) or "miranda meeting test" in n: return "Test / Internal"
    if has("gdpr","consent","marketing contact","unsubscrib","subscription","opt them into","opt-in","opt out"):
        return "Compliance & Marketing Status"
    if has("slack"): return "Slack & Notifications"
    if has("creates a note","pinned note","create note") or re.search(r"\bnote\b", n): return "Notes & Record Updates"
    if has("creates a task","create a task","call task","task to","creates a call"): return "Task & Call Creation"
    if has("assign","routing","round robin","rotate","forecast category","lead owner"):
        return "Lead Routing & Assignment"
    if has("lifecycle","lead score","scoring","became","sal date","stage based on","mql","sql"):
        return "Lifecycle & Scoring"
    if has("deal","pipeline","forecast","closed won","close date","rep confidence","customer success","cs-","onboarding survey","churn"):
        return "Deal & CS Ops"
    if has("email","follow-up","follow up","nurture","sequence","thank-you","thank you","reminder","send a series"):
        return "Email & Nurture"
    if has("copy","populate","enrich","clay","set the","sets the","portal id","property to","update contact data","data engine"):
        return "Data & Enrichment"
    return "Other / Uncategorized"

# ---- near-duplicate clustering (within object_type, on normalized names) -----
STOP = set("the and for to from of a an in on with new set copy flow workflow wf "
           "push slack email update updates create creates contact contacts company "
           "deal lead ff hubspot via based when".split())
def name_tokens(s):
    s = re.sub(r"\d+", " ", str(s or "").lower())
    s = re.sub(r"[^a-z\s]", " ", s)
    return set(w for w in s.split() if len(w) > 2 and w not in STOP)

def jaccard(a, b):
    if not a or not b: return 0.0
    i = len(a & b); return i / (len(a) + len(b) - i)

def cluster(rows, thresh=0.34):
    by_obj = defaultdict(list)
    for r in rows: by_obj[r["object_type"]].append(r)
    clusters = []
    for obj, items in by_obj.items():
        toks = [name_tokens(r["name"]) for r in items]
        used = set()
        for a in range(len(items)):
            if a in used: continue
            group = [items[a]]; used.add(a)
            for b in range(a+1, len(items)):
                if b in used: continue
                if jaccard(toks[a], toks[b]) >= thresh:
                    group.append(items[b]); used.add(b)
            if len(group) > 1:
                clusters.append((obj, group))
    return clusters

# ---- styling shortcuts ------------------------------------------------------
def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20

def widths(ws, w):
    for i, x in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = x

# ---- main -------------------------------------------------------------------
def build(csv_path, out_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    portal = re.search(r"_(\d+)\.csv$", os.path.basename(csv_path))
    portal = portal.group(1) if portal else "?"
    today = datetime.date.today().isoformat()

    for r in rows:
        r["_fn"] = classify(r)
    on  = [r for r in rows if is_on(r)]
    off = [r for r in rows if not is_on(r)]
    errored = [r for r in rows if num(r["active_issues"]) > 0]
    dormant = [r for r in on if num(r["enrolled_7d"]) == 0 and num(r["currently_enrolled"]) == 0]
    never   = [r for r in on if num(r["total_enrolled"]) == 0]
    dormant_only = [r for r in dormant if r not in never]
    clusters = cluster(rows)
    in_clusters = sum(len(g) for _, g in clusters)
    pct = round(100 * in_clusters / len(rows)) if rows else 0

    wb = openpyxl.Workbook()

    # ===== Diagnostic =====
    ws = wb.active; ws.title = "Diagnostic"; widths(ws, [4, 44, 12, 60])
    ws["B2"] = "HubSpot Workflow Audit — FullFunnel"; ws["B2"].font = Font(bold=True, size=16, color=DARK)
    ws["B3"] = f"Portal {portal} · {len(rows)} automation flows indexed · {today}"
    ws["B3"].font = Font(size=10, color=GREY)
    ws["B5"] = "Headline findings"; ws["B5"].font = Font(bold=True, size=12, color=ACCENT)
    findings = [
        ("Total automation flows", len(rows), "Indexed via internal CRM object 0-44"),
        ("Active (ON)", len(on), "Currently enabled"),
        ("Inactive (OFF)", len(off), "Archive / delete review"),
        ("Flows with HubSpot-flagged errors", len(errored), "Fix or retire — see Cleanup Queue"),
        ("Active but DORMANT", len(dormant), "On, but 0 enrolled in 7d and 0 currently in flow"),
        ("  …of those, NEVER enrolled anyone", len(never), "Strong delete candidates"),
        ("Near-duplicate clusters", len(clusters), "Consolidation opportunities — see Consolidation Map"),
        ("Workflows inside those clusters", in_clusters, f"{pct}% of all flows look like duplicates"),
    ]
    row = 6
    for label, val, note in findings:
        ws.cell(row=row, column=2, value=label)
        c = ws.cell(row=row, column=3, value=val); c.font = Font(bold=True); c.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4, value=note).font = Font(color=GREY, size=10)
        row += 1
    row += 1
    ws.cell(row=row, column=2, value="Recommended actions (priority order)").font = Font(bold=True, size=12, color=ACCENT); row += 1
    actions = [
        ("Fix or kill the flagged-error flows", "High", "Errors mean they may be silently failing. Triage in Cleanup Queue."),
        ("Retire never-enrolled active flows + review other dormant", "Medium", "On but idle = risk + clutter. Confirm not seasonal before deleting."),
        ("Consolidate the largest near-duplicate clusters", "High", "Highest-leverage dedup — see Consolidation Map."),
        ("Archive the OFF flows that are truly retired", "Low", "Reduces the in-app surface and the audit noise."),
        ("Establish naming + ownership governance", "Med", "Most clusters trace to one-off builds. Add an owner per flow."),
    ]
    hdr = ws.cell(row=row, column=2, value="Action"); hdr.font = Font(bold=True)
    ws.cell(row=row, column=3, value="Priority").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Why").font = Font(bold=True); row += 1
    for act, pri, why in actions:
        ws.cell(row=row, column=2, value=act); ws.cell(row=row, column=3, value=pri)
        ws.cell(row=row, column=4, value=why).font = Font(color=GREY, size=10); row += 1

    # ===== Consolidation Map =====
    ws = wb.create_sheet("Consolidation Map")
    cmh = ["#", "Object", "# flows", "On", "Off", "Member workflow", "Status", "Enr.7d"]
    ws.append(cmh); style_header(ws, 1, len(cmh)); widths(ws, [5, 16, 8, 6, 6, 46, 8, 9])
    clusters_sorted = sorted(clusters, key=lambda c: -len(c[1]))
    for i, (obj, group) in enumerate(clusters_sorted, 1):
        gon = sum(1 for r in group if is_on(r)); goff = len(group) - gon
        for j, r in enumerate(group):
            ws.append([i if j == 0 else "", obj.upper() if j == 0 else "",
                       len(group) if j == 0 else "", gon if j == 0 else "", goff if j == 0 else "",
                       r["name"], r["status"], num(r["enrolled_7d"])])
    ws.freeze_panes = "A2"

    # ===== Cleanup Queue =====
    ws = wb.create_sheet("Cleanup Queue")
    cqh = ["Priority", "Flag", "In dup cluster", "Workflow", "Status", "Object",
           "Function", "Enr.total", "Issues", "Updated", "Link"]
    ws.append(cqh); style_header(ws, 1, len(cqh)); widths(ws, [9, 22, 13, 42, 7, 14, 24, 10, 7, 12, 8])
    clustered_names = set(r["name"] for _, g in clusters for r in g)
    def queue_row(r, pri, flag):
        return [pri, flag, "Yes" if r["name"] in clustered_names else "", r["name"], r["status"],
                (r["object_type"] or "").upper(), r["_fn"], num(r["total_enrolled"]),
                num(r["active_issues"]), r["updated_on"], r["url"]]
    items = ([queue_row(r, "High", "Has errors") for r in errored] +
             [queue_row(r, "High", "Never enrolled anyone") for r in never] +
             [queue_row(r, "Medium", "Active but dormant") for r in dormant_only] +
             [queue_row(r, "Low", "Switched off") for r in off])
    for it in items:
        ws.append(it)
        link = it[-1]
        if link:
            cell = ws.cell(row=ws.max_row, column=len(cqh)); cell.value = "Open"; cell.hyperlink = link
            cell.font = Font(color=ACCENT, underline="single")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cqh))}{ws.max_row}"

    # ===== All / Active / Inactive =====
    disp = ["Workflow Name","Status","Object","Function","Trigger","Re-enroll","Description",
            "Enrolled (total)","Enrolled 7d","Unique","Currently","Issues","Created In",
            "Created","Created By","Updated","Updated By","Link","HubSpot ID"]
    def to_disp(r):
        return [r["name"], r["status"], (r["object_type"] or "").upper(), r["_fn"],
                trigger_label(r["trigger_type"]), r["reenrollment"], r["description"],
                num(r["total_enrolled"]), num(r["enrolled_7d"]), num(r["unique_enrolled"]),
                num(r["currently_enrolled"]), num(r["active_issues"]), r["created_in"],
                r["created_on"], r["created_by"], r["updated_on"], r["updated_by"],
                r["url"], r["id"]]
    def sheet(title, data):
        ws = wb.create_sheet(title)
        ws.append(disp); style_header(ws, 1, len(disp))
        widths(ws, [40,7,12,22,16,9,55,12,10,9,10,7,18,11,16,11,16,7,14])
        for r in data:
            ws.append(to_disp(r))
            link = r["url"]
            if link:
                cell = ws.cell(row=ws.max_row, column=18); cell.value = "Open"; cell.hyperlink = link
                cell.font = Font(color=ACCENT, underline="single")
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(disp))}{ws.max_row}"
    sheet("All Workflows", rows)
    sheet("Active (ON)", on)
    sheet("Inactive (OFF)", off)

    wb.save(out_path)
    return {"portal": portal, "rows": len(rows), "on": len(on), "off": len(off),
            "errored": len(errored), "dormant": len(dormant), "never": len(never),
            "clusters": len(clusters), "in_clusters": in_clusters, "out": out_path}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python3 build_audit.py <csv> [out.xlsx]"); sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else re.sub(r"\.csv$", ".xlsx", src)
    info = build(src, out)
    print("Built:", info["out"])
    print(f"  {info['rows']} flows | ON {info['on']} / OFF {info['off']} | "
          f"errored {info['errored']} | dormant {info['dormant']} (never {info['never']}) | "
          f"{info['clusters']} dup clusters ({info['in_clusters']} flows)")
